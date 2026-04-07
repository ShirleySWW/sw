import sys
from openpyxl import load_workbook

path = r"c:\Users\Yingxue.Wang\OneDrive - Laing ORourke\Fun\VS\sw\other\Procurement Register - SPW.xlsx"
sheet_name = "Procurement Schedule (Delivery)"

try:
    wb = load_workbook(path, read_only=True, data_only=True)
except Exception as e:
    print("ERROR loading workbook:", e)
    sys.exit(1)

if sheet_name not in wb.sheetnames:
    print("Sheet not found. Sheets:", wb.sheetnames)
    sys.exit(0)

ws = wb[sheet_name]

# Print header (first row)
for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
    headers = list(row)
    print("HEADERS:")
    for i, h in enumerate(headers, start=1):
        print(f"{i}: {h}")

# Print sample first 5 data rows
print("\nSAMPLE ROWS:")
count = 0
for row in ws.iter_rows(min_row=2, max_row=6, values_only=True):
    count += 1
    print(row)

if count == 0:
    print("No data rows found.")
